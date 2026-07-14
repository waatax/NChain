import re
from typing import List, Tuple, Dict, Any
from openpyxl.worksheet.worksheet import Worksheet
from .models import MnemonicItem, MnemonicItemSource, PairScene, SourceRef, NarrativeScene, StoryToken, NarrativeStory, Lesson
from .validators import normalize_number, ValidationIssue

def parse_canonical_items(ws: Worksheet) -> Tuple[List[MnemonicItem], List[ValidationIssue]]:
    items = []
    issues = []
    seen_numbers: Dict[str, str] = {}
    
    # Iterate through main table range (Rows 2-24, Columns B-O)
    for row_idx in range(2, 25):
        if row_idx in (13, 14):
            continue
        row = ws[row_idx]
        # We only check columns B (2), E (5), H (8), K (11), N (14)
        for col_idx in (2, 5, 8, 11, 14):
            cell = row[col_idx - 1] # 0-indexed list
            val = cell.value
            if val is None:
                continue
            
            val_str = str(val).strip()
            if not val_str.isdigit():
                continue
                
            try:
                num = normalize_number(val_str)
            except ValueError:
                continue
                
            # Right neighbor contains the keyword
            keyword_cell = row[col_idx] # next column is at index col_idx (since col_idx is 1-based index)
            keyword_val = keyword_cell.value
            if keyword_val is not None:
                keyword = str(keyword_val).strip()
                if not keyword:
                    continue
                
                cell_ref = f"{cell.coordinate}"
                if num in seen_numbers:
                    issues.append(ValidationIssue(
                        severity="error",
                        code="DUPLICATE_ITEM",
                        message=f"Duplicate canonical mapping for number {num} in cell {cell_ref}. Previously seen in {seen_numbers[num]}.",
                        source={"sheet": ws.title, "cell": cell_ref},
                        context={"expected": seen_numbers[num], "actual": cell_ref}
                    ))
                else:
                    seen_numbers[num] = cell_ref
                    items.append(MnemonicItem(
                        id=f"item-{num}",
                        number=num,
                        numericValue=int(num),
                        canonicalKeyword=keyword,
                        aliases=[],
                        source=MnemonicItemSource(sheet=ws.title, cell=cell_ref)
                    ))
                        
    # Check if all 101 items are found
    for i in range(101):
        num_str = "100" if i == 100 else f"{i:02d}"
        if num_str not in seen_numbers:
            issues.append(ValidationIssue(
                severity="error",
                code="MISSING_ITEM",
                message=f"Missing canonical mapping for number {num_str} in sheet {ws.title}.",
                source={"sheet": ws.title, "cell": "N/A"}
            ))
            
    return items, issues

def parse_pairwise_scenes(ws: Worksheet) -> Tuple[List[PairScene], List[ValidationIssue]]:
    scenes = []
    issues = []
    
    # Matches: 【00 鎖鏈】 ➔ 【01 葉子】 or similar arrows like →, ➡
    title_re = re.compile(r"【\s*(\d{1,3})\s+([^】]+?)\s*】\s*[➔→➡]\s*【\s*(\d{1,3})\s+([^】]+?)\s*】")
    scene_re = re.compile(r"^\s*畫面\s*[：:]\s*(.+?)\s*$")
    
    current_title = None
    current_title_cell = None
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
        for col_idx, cell in enumerate(row, start=1):
            val = cell.value
            if val is None or not isinstance(val, str):
                continue
            
            title_match = title_re.search(val)
            if title_match:
                if current_title:
                    issues.append(ValidationIssue(
                        severity="error",
                        code="PARSE_FAILURE",
                        message=f"Found title '{val}' before a matching scene text for the previous title '{current_title[0]}->{current_title[2]}'.",
                        source={"sheet": ws.title, "cell": f"{cell.coordinate}"}
                    ))
                current_title = (
                    title_match.group(1),
                    title_match.group(2).strip(),
                    title_match.group(3),
                    title_match.group(4).strip()
                )
                current_title_cell = cell
                continue
                
            scene_match = scene_re.search(val)
            if scene_match and current_title:
                scene_text = scene_match.group(1).strip()
                
                try:
                    from_num = normalize_number(current_title[0])
                    to_num = normalize_number(current_title[2])
                except ValueError as e:
                    issues.append(ValidationIssue(
                        severity="error",
                        code="PARSE_FAILURE",
                        message=f"Failed to normalize numbers in title: {str(e)}",
                        source={"sheet": ws.title, "cell": f"{current_title_cell.coordinate}"}
                    ))
                    current_title = None
                    continue
                
                from_val = int(from_num)
                to_val = int(to_num)
                
                if to_val - from_val != 1:
                    issues.append(ValidationIssue(
                        severity="error",
                        code="RANGE_MISMATCH",
                        message=f"Invalid pair sequence: from {from_num} to {to_num} (must be sequential with gap of 1)",
                        source={"sheet": ws.title, "cell": f"{current_title_cell.coordinate}"}
                    ))
                
                # Determine lesson ID based on range
                lesson_id = None
                if 0 <= from_val <= 9:
                    lesson_id = "lesson-00-10"
                elif 11 <= from_val <= 19:
                    lesson_id = "lesson-11-20"
                elif 21 <= from_val <= 29:
                    lesson_id = "lesson-21-30"
                elif 31 <= from_val <= 39:
                    lesson_id = "lesson-31-40"
                elif 41 <= from_val <= 49:
                    lesson_id = "lesson-41-50"
                elif 51 <= from_val <= 59:
                    lesson_id = "lesson-51-60"
                else:
                    issues.append(ValidationIssue(
                        severity="error",
                        code="RANGE_MISMATCH",
                        message=f"Pair {from_num}->{to_num} falls outside of pairwise lesson ranges (00-10, 11-20, 21-30, 31-40, 41-50, 51-60).",
                        source={"sheet": ws.title, "cell": f"{current_title_cell.coordinate}"}
                    ))
                    lesson_id = "lesson-unknown"
                
                if "?" in val or "？" in val:
                    issues.append(ValidationIssue(
                        severity="warning",
                        code="DAMAGED_CHAR",
                        message=f"Damaged character '?' found in scene text: {val}",
                        source={"sheet": ws.title, "cell": f"{cell.coordinate}"}
                    ))
                
                scenes.append(PairScene(
                    id=f"pair-{from_num}-{to_num}",
                    kind="pair",
                    lessonId=lesson_id,
                    order=len(scenes) + 1,
                    fromItemId=f"item-{from_num}",
                    toItemId=f"item-{to_num}",
                    displayFromKeyword=current_title[1],
                    displayToKeyword=current_title[3],
                    sceneText=scene_text,
                    source=SourceRef(
                        sheet=ws.title,
                        rowStart=current_title_cell.row,
                        rowEnd=cell.row,
                        rawText=f"{current_title_cell.value} | {cell.value}"
                    )
                ))
                current_title = None
                current_title_cell = None
                
    if current_title:
        issues.append(ValidationIssue(
            severity="error",
            code="PARSE_FAILURE",
            message=f"Unmatched title at end of sheet '{current_title[0]}->{current_title[2]}'",
            source={"sheet": ws.title, "cell": f"{current_title_cell.coordinate}"}
        ))
        
    return scenes, issues

def tokenize_sentence(text: str) -> Tuple[List[StoryToken], List[str]]:
    tokens = []
    item_ids = []
    # Token matches (61) 老人 or (62)驢兒
    token_pattern = re.compile(r"\((?P<number>\d{1,3})\)\s*(?P<keyword>[^，。！、；：()\s]+?)(?=[，。！、；：()]|$)")
    
    last_idx = 0
    for match in token_pattern.finditer(text):
        start, end = match.span()
        if start > last_idx:
            # Add plain text segment before match
            tokens.append(StoryToken(text=text[last_idx:start]))
        
        num_str = normalize_number(match.group("number"))
        item_id = f"item-{num_str}"
        item_ids.append(item_id)
        
        tokens.append(StoryToken(text=match.group(0), itemId=item_id))
        last_idx = end
        
    if last_idx < len(text):
        tokens.append(StoryToken(text=text[last_idx:]))
        
    return tokens, item_ids

def parse_narrative_stories(ws: Worksheet) -> Tuple[List[NarrativeStory], List[NarrativeScene], List[ValidationIssue]]:
    stories = []
    scenes = []
    issues = []
    
    title_pattern = re.compile(r"🧠\s*數字鎖鏈記憶故事：《(?P<title>[^》]+)》")
    token_pattern = re.compile(r"\((?P<number>\d{1,3})\)")
    
    # First, let's partition the rows of the sheet into story blocks
    story_blocks: List[Dict[str, Any]] = []
    current_block = None
    
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # We look at Col1 (index 0) or Col2 (index 1) for the title
        cell_val = None
        for val in row[:2]:
            if val is not None and isinstance(val, str):
                cell_val = val
                break
                
        if cell_val:
            title_match = title_pattern.search(cell_val)
            if title_match:
                if current_block:
                    story_blocks.append(current_block)
                current_block = {
                    "title": title_match.group("title"),
                    "start_row": row_idx,
                    "rows": [],
                    "title_coordinate": f"A{row_idx}"
                }
                continue
                
        if current_block:
            current_block["rows"].append((row_idx, row))
            
    if current_block:
        story_blocks.append(current_block)
        
    # Process each story block
    for block in story_blocks:
        title = block["title"]
        block_scenes: List[NarrativeScene] = []
        memory_tip_parts = []
        recap_parts = []
        
        # Determine lesson ID based on title or content
        # L07: 61-70, L08: 71-80, L09: 81-90, L10: 91-100
        lesson_id = "lesson-unknown"
        if "農莊" in title or "61" in title or "80" in title and "極地" not in title:
            # wait, 61-80 has agricultural adventure ("農莊冒險" for 61-70) and polar adventure ("極地與巴黎的奇幻漂流" for 71-80)
            if "農莊" in title:
                lesson_id = "lesson-61-70"
            elif "極地" in title:
                lesson_id = "lesson-71-80"
        elif "西域" in title or "81" in title:
            lesson_id = "lesson-81-90"
        elif "操場" in title or "91" in title:
            lesson_id = "lesson-91-100"
            
        is_tip = False
        is_recap = False
        
        for row_idx, row in block["rows"]:
            # Find first non-empty cell in row
            cell_val = None
            cell_col = 0
            for col_i, val in enumerate(row):
                if val is not None and isinstance(val, str):
                    cell_val = val.strip()
                    cell_col = col_i + 1
                    break
            
            if not cell_val:
                continue
                
            cell_coordinate = f"{ws.cell(row=row_idx, column=cell_col).coordinate}"
            
            if cell_val == "↓":
                continue
                
            # If we hit memory tip header
            if "記憶小秘訣" in cell_val or "記憶小技巧" in cell_val or "回想練習" in cell_val or "回顧小秘訣" in cell_val:
                is_tip = True
                is_recap = False
                continue
                
            # If we hit recap header
            if "複習專用" in cell_val or "記憶聯想關鍵點" in cell_val:
                is_recap = True
                is_tip = False
                continue
                
            if is_tip:
                memory_tip_parts.append(cell_val)
                continue
                
            if is_recap:
                # We skip table headers or details, but let's capture it if it's normal text
                if not cell_val.startswith("數字") and not cell_val.isdigit():
                    recap_parts.append(cell_val)
                continue
                
            # Check if this row contains parenthesized numbers, which makes it a scene sentence
            if token_pattern.search(cell_val):
                if "?" in cell_val or "？" in cell_val:
                    issues.append(ValidationIssue(
                        severity="warning",
                        code="DAMAGED_CHAR",
                        message=f"Damaged character '?' found in narrative scene: {cell_val}",
                        source={"sheet": ws.title, "cell": cell_coordinate}
                    ))
                
                # Tokenize
                tokens, item_ids = tokenize_sentence(cell_val)
                
                scene_id = f"story-{lesson_id[7:]}-scene-{len(block_scenes) + 1:02d}"
                block_scenes.append(NarrativeScene(
                    id=scene_id,
                    kind="narrative-scene",
                    storyId=f"story-{lesson_id[7:]}",
                    lessonId=lesson_id,
                    order=len(block_scenes) + 1,
                    originalText=cell_val,
                    itemIds=item_ids,
                    tokens=tokens,
                    source=SourceRef(
                        sheet=ws.title,
                        rowStart=row_idx,
                        rowEnd=row_idx,
                        rawText=cell_val
                    )
                ))
                
        # Aggregate memory tip and recap
        memory_tip = "\n".join(memory_tip_parts).strip() if memory_tip_parts else None
        recap_text = "\n".join(recap_parts).strip() if recap_parts else None
        
        story_id = f"story-{lesson_id[7:]}"
        stories.append(NarrativeStory(
            id=story_id,
            lessonId=lesson_id,
            title=title,
            sceneIds=[s.id for s in block_scenes],
            recapText=recap_text,
            memoryTip=memory_tip
        ))
        scenes.extend(block_scenes)
        
    return stories, scenes, issues
