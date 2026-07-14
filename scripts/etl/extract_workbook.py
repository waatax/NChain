import os
import sys
import json
import hashlib
import argparse
from datetime import datetime
from openpyxl import load_workbook
from typing import List, Dict, Any

# Adjust path to import local modules
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from etl.models import (
    MnemonicItem, PairScene, NarrativeScene, NarrativeStory, Lesson,
    ContentManifest, ManifestCounts, SourceRef
)
from etl.validators import normalize_number, ValidationIssue
from etl.parsers import parse_canonical_items, parse_pairwise_scenes, parse_narrative_stories

def compute_sha256(filepath: str) -> str:
    h = hashlib.sha256()
    with open(filepath, 'rb') as f:
        for chunk in iter(lambda: f.read(65536), b''):
            h.update(chunk)
    return h.hexdigest()

def main():
    parser = argparse.ArgumentParser(description="ETL script for NChain spreadsheet parsing")
    parser.add_argument("--input", required=True, help="Path to input Excel file")
    parser.add_argument("--output", required=True, help="Directory to save JSON files")
    parser.add_argument("--report", required=True, help="Path to write JSON validation report")
    parser.add_argument("--strict", action="store_true", help="Fail and exit 1 on any validation errors")
    args = parser.parse_args()
    
    if not os.path.exists(args.input):
        print(f"Error: Input file {args.input} does not exist", file=sys.stderr)
        sys.exit(2)
        
    os.makedirs(args.output, exist_ok=True)
    os.makedirs(os.path.dirname(args.report), exist_ok=True)
    
    print(f"Loading workbook: {args.input}")
    wb = load_workbook(args.input, data_only=True)
    
    issues: List[ValidationIssue] = []
    
    # 1. Parse canonical items
    print("Parsing canonical items from sheet '1-100'...")
    if "1-100" not in wb.sheetnames:
        print("Error: Missing canonical sheet '1-100'", file=sys.stderr)
        sys.exit(2)
        
    items, item_issues = parse_canonical_items(wb["1-100"])
    issues.extend(item_issues)
    
    # Map items by numericValue for fast lookup
    items_dict: Dict[int, MnemonicItem] = {item.numericValue: item for item in items}
    
    # 2. Parse pairwise sheets
    pairwise_sheets = ["00-20", "21-40", "41-60"]
    pair_scenes: List[PairScene] = []
    for sheet_name in pairwise_sheets:
        if sheet_name in wb.sheetnames:
            print(f"Parsing pairwise scenes from sheet '{sheet_name}'...")
            scenes, sheet_issues = parse_pairwise_scenes(wb[sheet_name])
            pair_scenes.extend(scenes)
            issues.extend(sheet_issues)
        else:
            issues.append(ValidationIssue(
                severity="error",
                code="MISSING_ITEM",
                message=f"Missing expected pairwise sheet: {sheet_name}"
            ))
            
    # 3. Parse narrative sheets
    narrative_sheets = ["61-80", "81-100"]
    stories: List[NarrativeStory] = []
    narrative_scenes: List[NarrativeScene] = []
    for sheet_name in narrative_sheets:
        if sheet_name in wb.sheetnames:
            print(f"Parsing narrative stories from sheet '{sheet_name}'...")
            sheet_stories, sheet_scenes, sheet_issues = parse_narrative_stories(wb[sheet_name])
            stories.extend(sheet_stories)
            narrative_scenes.extend(sheet_scenes)
            issues.extend(sheet_issues)
        else:
            issues.append(ValidationIssue(
                severity="error",
                code="MISSING_ITEM",
                message=f"Missing expected narrative sheet: {sheet_name}"
            ))
            
    # Post-process: Extract display keyword aliases and warning validation checks
    print("Running post-process validation and alias mapping...")
    
    # Map from display keywords to items
    # Check pairwise scenes
    for p_scene in pair_scenes:
        # Check fromItem display keyword vs canonical
        from_val = int(p_scene.fromItemId.split("-")[1])
        to_val = int(p_scene.toItemId.split("-")[1])
        
        from_item = items_dict.get(from_val)
        to_item = items_dict.get(to_val)
        
        if from_item:
            disp = p_scene.displayFromKeyword
            canonical = from_item.canonicalKeyword
            if disp != canonical:
                issues.append(ValidationIssue(
                    severity="warning",
                    code="KEYWORD_VARIANT",
                    message=f"Display variant for item-{from_val:02d}: display '{disp}' vs canonical '{canonical}'",
                    source={"sheet": p_scene.source.sheet, "cell": f"Row {p_scene.source.rowStart}"},
                    context={"expected": canonical, "actual": disp}
                ))
                if disp not in from_item.aliases:
                    from_item.aliases.append(disp)
                    
        if to_item:
            disp = p_scene.displayToKeyword
            canonical = to_item.canonicalKeyword
            if disp != canonical:
                issues.append(ValidationIssue(
                    severity="warning",
                    code="KEYWORD_VARIANT",
                    message=f"Display variant for item-{to_val:02d}: display '{disp}' vs canonical '{canonical}'",
                    source={"sheet": p_scene.source.sheet, "cell": f"Row {p_scene.source.rowStart}"},
                    context={"expected": canonical, "actual": disp}
                ))
                if disp not in to_item.aliases:
                    to_item.aliases.append(disp)
                    
    # Check narrative scenes
    for n_scene in narrative_scenes:
        # In narrative scene tokens, map any custom spelling to aliases
        for token in n_scene.tokens:
            if token.itemId:
                item_val = int(token.itemId.split("-")[1])
                item = items_dict.get(item_val)
                if item:
                    # Token format is e.g. "(61) 老人"
                    # The keyword is the part after the paren
                    clean_kw = token.text.split(")")[-1].strip()
                    canonical = item.canonicalKeyword
                    if clean_kw != canonical:
                        issues.append(ValidationIssue(
                            severity="warning",
                            code="KEYWORD_VARIANT",
                            message=f"Display variant for item-{item_val:02d}: display '{clean_kw}' vs canonical '{canonical}'",
                            source={"sheet": n_scene.source.sheet, "cell": f"Row {n_scene.source.rowStart}"},
                            context={"expected": canonical, "actual": clean_kw}
                        ))
                        if clean_kw not in item.aliases:
                            item.aliases.append(clean_kw)

    # 4. Generate Lessons
    print("Generating lessons...")
    lessons: List[Lesson] = []
    
    lesson_ranges = [
        ("lesson-00-10", "module-00-20", "00", "10", "pair", 1),
        ("lesson-11-20", "module-00-20", "11", "20", "pair", 2),
        ("lesson-21-30", "module-21-40", "21", "30", "pair", 3),
        ("lesson-31-40", "module-21-40", "31", "40", "pair", 4),
        ("lesson-41-50", "module-41-60", "41", "50", "pair", 5),
        ("lesson-51-60", "module-41-60", "51", "60", "pair", 6),
        ("lesson-61-70", "module-61-80", "61", "70", "narrative", 7),
        ("lesson-71-80", "module-61-80", "71", "80", "narrative", 8),
        ("lesson-81-90", "module-81-100", "81", "90", "narrative", 9),
        ("lesson-91-100", "module-81-100", "91", "100", "narrative", 10),
    ]
    
    lesson_titles = {
        "lesson-00-10": "00–10 鎖鏈至汽水",
        "lesson-11-20": "11–20 筷子至風鈴",
        "lesson-21-30": "21–30 鱷魚至別針",
        "lesson-31-40": "31–40 山衣至司令",
        "lesson-41-50": "41–50 死魚至警長",
        "lesson-51-60": "51–60 引擎至手槍",
        "lesson-61-70": "61–70 瘋狂的農莊冒險",
        "lesson-71-80": "71–80 極地與巴黎的奇幻冒險",
        "lesson-81-90": "81–90 西域巴士與微醺精靈",
        "lesson-91-100": "91–100 從操場到酒吧的瘋狂派對",
    }
    
    for l_id, m_id, start_str, end_str, mode, order in lesson_ranges:
        start_val = int(start_str)
        end_val = int(end_str)
        
        # Collect item IDs
        l_item_ids = [f"item-{i:02d}" if i < 100 else "item-100" for i in range(start_val, end_val + 1)]
        
        # Collect scene IDs
        if mode == "pair":
            l_scene_ids = [s.id for s in pair_scenes if s.lessonId == l_id]
        else:
            l_scene_ids = [s.id for s in narrative_scenes if s.lessonId == l_id]
            
        lessons.append(Lesson(
            id=l_id,
            moduleId=m_id,
            title=lesson_titles[l_id],
            rangeStart=start_str,
            rangeEnd=end_str,
            mode=mode,
            itemIds=l_item_ids,
            sceneIds=l_scene_ids,
            order=order
        ))
        
    # Generate Modules metadata list
    modules = [
        {"id": "module-00-20", "title": "數字 00–20 基礎與日常聯想", "lessonIds": ["lesson-00-10", "lesson-11-20"], "order": 1},
        {"id": "module-21-40", "title": "數字 21–40 進階物件聯想", "lessonIds": ["lesson-21-30", "lesson-31-40"], "order": 2},
        {"id": "module-41-60", "title": "數字 41–60 複雜場景配對", "lessonIds": ["lesson-41-50", "lesson-51-60"], "order": 3},
        {"id": "module-61-80", "title": "數字 61–80 劇本式連鎖記憶", "lessonIds": ["lesson-61-70", "lesson-71-80"], "order": 4},
        {"id": "module-81-100", "title": "數字 81–100 終極長編故事記憶", "lessonIds": ["lesson-81-90", "lesson-91-100"], "order": 5},
    ]

    # Check for strict errors
    errors_count = sum(1 for issue in issues if issue.severity == "error")
    warnings_count = sum(1 for issue in issues if issue.severity == "warning")
    
    print(f"ETL run finished. Validation results: {errors_count} errors, {warnings_count} warnings.")
    
    # Save validation issues report
    with open(args.report, 'w', encoding='utf-8') as f:
        json.dump([issue.model_dump() for issue in issues], f, ensure_ascii=False, indent=2)
    print(f"Validation report saved to {args.report}")
    
    # Strict mode termination
    if args.strict and errors_count > 0:
        print("Error: Strict mode enabled and validation errors were found. Aborting JSON generation.", file=sys.stderr)
        sys.exit(1)
        
    # Write JSON files to output directory
    print(f"Writing JSON data outputs to: {args.output}")
    
    # Write items.json
    with open(os.path.join(args.output, "items.json"), 'w', encoding='utf-8') as f:
        json.dump([item.model_dump() for item in items], f, ensure_ascii=False, indent=2)
        
    # Write lessons.json
    with open(os.path.join(args.output, "lessons.json"), 'w', encoding='utf-8') as f:
        json.dump([l.model_dump() for l in lessons], f, ensure_ascii=False, indent=2)
        
    # Write modules.json
    with open(os.path.join(args.output, "modules.json"), 'w', encoding='utf-8') as f:
        json.dump(modules, f, ensure_ascii=False, indent=2)
        
    # Write pair-scenes.json
    with open(os.path.join(args.output, "pair-scenes.json"), 'w', encoding='utf-8') as f:
        json.dump([s.model_dump() for s in pair_scenes], f, ensure_ascii=False, indent=2)
        
    # Write stories.json
    with open(os.path.join(args.output, "stories.json"), 'w', encoding='utf-8') as f:
        json.dump([s.model_dump() for s in stories], f, ensure_ascii=False, indent=2)
        
    # Write narrative-scenes.json
    with open(os.path.join(args.output, "narrative-scenes.json"), 'w', encoding='utf-8') as f:
        json.dump([s.model_dump() for s in narrative_scenes], f, ensure_ascii=False, indent=2)
        
    # Write content-index.json
    content_index = {
        "items": {item.id: "items.json" for item in items},
        "lessons": {l.id: "lessons.json" for l in lessons},
        "stories": {s.id: "stories.json" for s in stories},
        "pairScenes": {s.id: "pair-scenes.json" for s in pair_scenes},
        "narrativeScenes": {s.id: "narrative-scenes.json" for s in narrative_scenes}
    }
    with open(os.path.join(args.output, "content-index.json"), 'w', encoding='utf-8') as f:
        json.dump(content_index, f, ensure_ascii=False, indent=2)
        
    # Generate content version manifest
    source_sha = compute_sha256(args.input)
    manifest = ContentManifest(
        contentVersion=datetime.now().strftime("%Y.%m.%d.%H"),
        generatedAt=datetime.utcnow().isoformat() + "Z",
        sourceFileSha256=source_sha,
        counts=ManifestCounts(
            items=len(items),
            lessons=len(lessons),
            pairScenes=len(pair_scenes),
            stories=len(stories),
            narrativeScenes=len(narrative_scenes)
        )
    )
    
    with open(os.path.join(args.output, "manifest.json"), 'w', encoding='utf-8') as f:
        json.dump(manifest.model_dump(), f, ensure_ascii=False, indent=2)
        
    # Also write a readable Markdown report
    md_report_path = os.path.splitext(args.report)[0] + ".md"
    with open(md_report_path, 'w', encoding='utf-8') as f:
        f.write("# ETL Extraction Validation Report\n\n")
        f.write(f"**Date**: {manifest.generatedAt}\n")
        f.write(f"**Source File SHA256**: `{source_sha}`\n\n")
        f.write("## Entity Counts\n")
        f.write(f"- **Canonical Items (00-100)**: {len(items)}\n")
        f.write(f"- **Lessons**: {len(lessons)}\n")
        f.write(f"- **Pairwise Scenes**: {len(pair_scenes)}\n")
        f.write(f"- **Narrative Stories**: {len(stories)}\n")
        f.write(f"- **Narrative Scenes**: {len(narrative_scenes)}\n\n")
        f.write("## Validation Issues\n")
        if not issues:
            f.write("✅ No issues found. Data is healthy.\n")
        else:
            f.write(f"Found {errors_count} errors and {warnings_count} warnings.\n\n")
            f.write("| Severity | Code | Message | Location |\n")
            f.write("| --- | --- | --- | --- |\n")
            for issue in issues:
                loc = f"{issue.source.get('sheet', '')} {issue.source.get('cell', '')}" if issue.source else "N/A"
                f.write(f"| {issue.severity.upper()} | {issue.code} | {issue.message} | {loc} |\n")
                
    print(f"Readable validation report saved to {md_report_path}")

if __name__ == "__main__":
    main()
